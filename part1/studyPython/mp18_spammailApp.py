# 대량 메일 전송 
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

workbook = load_workbook('./part1/studyPython/spamMailList.xlsx')
worksheet = workbook.active # sheet1 선택

for i in range(1,worksheet.max_row + 1):
    recv_email = worksheet.cell(i,1).value
    print(recv_email)
    try:
        #실제 메일 전송 로직
        send_email = 'hanbv@naver.com'
        send_pass = '----'
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        msg = MIMEMultipart()
        msg['subject'] = '엑셀에서 보내는 메일'
        msg['From'] = send_email
        msg['To'] = recv_email
        msg.attach(MIMEText('보내는 내용입니다. 메롱~!!'))

        mail = smtplib.SMTP(smtp_name, smtp_port)
        mail.starttls()
        mail.login(send_email,send_pass)
        mail.sendmail(send_email,recv_email, msg.as_string())
        mail.quit()
        print(f'전송성공 : {recv_email}')

    except Exception as e:
        print(f'수신메일 - {recv_email}')
        print(f'전송에러 : {e}')